"""
GNOMON EXPERTS - Εφαρμογή Εκθέσεων v2.1
pip3 install streamlit openpyxl Pillow beautifulsoup4
streamlit run app.py
"""

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
import io, subprocess, tempfile, os, re
from datetime import datetime
from PIL import Image as PILImage

# Database
from vehicles import get_markes, get_montela
from pdf_generator import generate_pdf
from database import (init_db, save_ekthesi, load_ekthesi, search_ektheseis,
                      delete_ekthesi, get_statistics, update_status,
                      get_history, export_all_json,
                      get_axia_stats, get_antallaktiko_stats_all_types,
    get_synergeio_eponimies, get_synergeio_full,
    get_custom_markes, get_custom_montela, add_custom_marka,
    add_custom_montelo, delete_custom_vehicle)

TEMPLATE_FILE = "ekthesi_clean.xlsx"

PARATIRISI_TEMPLATES = {
    "1. Εμπορική αξία οχήματος": (
        "Για τον προσδιορισμό της εμπορικής αξίας του προαναφερόμενου οχήματος "
        "λήφθηκαν υπ' όψιν οι μέγιστες & ελάχιστες τιμές εμπορικής αξίας "
        "αντίστοιχων οχημάτων στην Ελλάδα με αποτέλεσμα τον προσδιορισμό της "
        "μέσης ενδεικτικής τιμής εμπορικής αξίας του συγκεκριμένου οχήματος."
    ),
    "2. Αρχική εκτίμηση (πριν αποσυναρμολόγηση)": (
        "Η παραπάνω αρχική εκτίμηση της συνολικής δαπάνης αποκατάστασης των Υ.Ζ "
        "του ζημιωθέντος Ε.Ι.Χ, είναι ενδεικτική και έγινε χωρίς να προχωρήσουν "
        "οι εργασίες αποσυναρμολόγησής του. Το μέγεθος της δαπάνης πιθανόν να "
        "διαφοροποιηθεί, εάν και εφόσον προχωρήσουν οι εργασίες αποσυναρμολόγησης "
        "και επισκευής του. Το αρχικό αναγραφόμενο κόστος που παρουσιάζεται είναι "
        "μη συμφωνημένο με τον επισκευαστή και αφορά κατ' εκτίμηση τις αρχικές "
        "εργασίες πριν την αποσυναρμολόγηση του οχήματος."
    ),
    "3. Τελική εκτίμηση (συμφωνημένη)": (
        "Η ανωτέρω πραγματογνωμοσύνη είναι τελική & ολοκληρωμένη. "
        "Το κόστος εργασιών (φανοποιεία & βαφή) έχει συμφωνηθεί με τον "
        "επισκευαστή στα: ___ ευρώ συν ΦΠΑ.\n"
        "Την τελική συμφωνία αποζημίωσης του κόστους εργασιών & ανταλλακτικών "
        "της ανωτέρω πραγματογνωμοσύνης, διαχειρίζεται η ασφαλιστική εταιρεία "
        "η οποία έχει & την ευθύνη του διακανονισμού."
    ),
}

# ============================================================
# HTML PARSING
# ============================================================
def parse_html_report(html_content):
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return None, "Χρειάζεται: pip3 install beautifulsoup4"
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        data = {}

        def get_val(label):
            for b in soup.find_all('b'):
                if b.get_text(strip=True) == label:
                    td = b.find_parent('td')
                    if td:
                        nxt = td.find_next_sibling('td')
                        if nxt:
                            return nxt.get_text(strip=True)
            return ""

        data['ar_zimias']      = get_val('Αριθμ. Ζημίας:')
        data['hm_entolhs']     = get_val('Ημ/νία εντολής:')
        hm = get_val('Ημ/νία ατυχήματος:')
        data['hm_atyxhmatos']  = hm.split()[0] if hm else ''
        data['idioktitis']     = get_val('Ιδιοκτήτης:')
        data['xrisi']          = get_val('Χρήση:')
        data['ar_kykloforias'] = get_val('Αριθμ. Κυκλοφορίας:')
        data['montelo']        = get_val('Μοντέλο:')
        data['marka']          = get_val('Μάρκα:')
        data['kyvika']         = get_val('Κυβικά:')
        data['proti_adeia']    = get_val('1η Άδεια Κυκλοφορίας:')
        data['xiliometrites']  = get_val('Ένδειξη Χιλιομετρητή:')
        axia_raw               = get_val('Αξία:')
        data['axia']           = axia_raw.replace('€','').replace('.','').replace(',','').strip()
        data['visit_date']     = get_val('Ημ/νία 1ης Επίσκεψης:')
        data['visit_place']    = get_val('Τόπος:')

        topos = []
        for b in soup.find_all('b'):
            if b.get_text(strip=True) == 'Διεύθυνση:':
                v = b.next_sibling
                if v: topos.append(str(v).strip())
            elif b.get_text(strip=True) == 'Περιοχή:':
                v = b.next_sibling
                if v: topos.append(str(v).strip())
        data['topos_atyxhmatos'] = ' '.join(topos)

        parts, works = [], []
        target_table = None
        for tbl in soup.find_all('table'):
            if any('Περιγραφή' in td.get_text() for td in tbl.find_all('td', class_='tl')):
                target_table = tbl
                break

        if target_table:
            for row in target_table.find_all('tr'):
                cells = row.find_all('td')
                if len(cells) < 2: continue
                desc = cells[0].get_text(strip=True)
                if not desc: continue
                if any(s in desc for s in ['Σύνολα','Φ.Π.Α','Γενικό','Περιγραφή']): continue

                def pv(c):
                    try: return float(c.get_text(strip=True).replace('.','').replace(',','.'))
                    except: return 0.0

                axia_v  = pv(cells[1]) if len(cells)>1 else 0.0
                fanop_v = pv(cells[2]) if len(cells)>2 else 0.0
                vaf_v   = pv(cells[3]) if len(cells)>3 else 0.0
                mix_v   = pv(cells[4]) if len(cells)>4 else 0.0

                ptype = ''
                clean = desc
                m = re.search(r'\(([^)]*)\)\s*$', desc)
                if m:
                    ptype = {'Γν.':'ΓΝ','ΓΝ':'ΓΝ','ΜΤΧ':'ΜΤΧ','ΙΜ':'ΙΜ'}.get(m.group(1).strip(),'')
                    clean = desc[:m.start()].strip()

                du = desc.upper()
                if 'ΑΝΤΙΚΑΤΑΣΤΑΣΗ' in du:
                    name = re.sub(r'^ΑΝΤΙΚΑΤΑΣΤΑΣΗ\s*','',clean,flags=re.IGNORECASE).strip()
                    parts.append({'name':name,'type':ptype,'price':axia_v,'fanop':fanop_v,'vafeas':vaf_v,'mixanikos':mix_v,'ilgos':0.0})
                elif 'ΕΞΑΓΩΓΗ' in du or 'ΤΟΠΟΘΕΤΗΣΗ' in du:
                    wd = re.sub(r'^ΕΞΑΓΩΓΗ[\-\s]ΤΟΠΟΘΕΤΗΣΗ\s*','',clean,flags=re.IGNORECASE).strip()
                    works.append({'type':'ΕΞΑΓΩΓΗ-ΤΟΠΟΘΕΤΗΣΗ','desc':wd,'fanop':fanop_v,'vafeas':vaf_v,'mixanikos':mix_v,'ilgos':0.0})
                elif 'ΕΠΙΣΚΕΥΗ' in du:
                    wd = re.sub(r'^ΕΠΙΣΚΕΥΗ\s*','',clean,flags=re.IGNORECASE).strip()
                    works.append({'type':'ΕΠΙΣΚΕΥΗ','desc':wd,'fanop':fanop_v,'vafeas':vaf_v,'mixanikos':mix_v,'ilgos':0.0})
                else:
                    works.append({'type':'ΕΠΙΣΚΕΥΗ','desc':clean,'fanop':fanop_v,'vafeas':vaf_v,'mixanikos':mix_v,'ilgos':0.0})

        data['parts'] = parts
        data['works'] = works
        return data, None
    except Exception as e:
        return None, str(e)


def apply_to_session(data):
    """Γράφει τα parsed δεδομένα απευθείας στα session_state keys των widgets."""
    st.session_state['ar_zimias']        = data.get('ar_zimias','')
    st.session_state['topos_atyxhmatos'] = data.get('topos_atyxhmatos','')
    st.session_state['idioktitis']       = data.get('idioktitis','')
    st.session_state['ar_kykloforias']   = data.get('ar_kykloforias','')
    st.session_state['marka']            = data.get('marka','')
    st.session_state['proti_adeia']      = data.get('proti_adeia','')
    st.session_state['xiliometrites']    = data.get('xiliometrites','')
    st.session_state['xrisi']            = data.get('xrisi','')
    st.session_state['kyvika']           = data.get('kyvika','')
    st.session_state['montelo']          = data.get('montelo','')
    st.session_state['vd_text_0']        = data.get('visit_date','')
    st.session_state['visit_place_0']    = data.get('visit_place','')
    try:
        st.session_state['axia'] = int(float(data.get('axia',0))) if data.get('axia') else 0
    except:
        st.session_state['axia'] = 0

    parts_data = data.get('parts',[])
    st.session_state['num_parts'] = max(len(parts_data), 1)
    PTYPS = ["","ΓΝ","ΜΤΧ","ΙΜ"]
    for i, p in enumerate(parts_data):
        st.session_state[f'p_name_{i}']   = p.get('name','')
        pt = p.get('type','')
        st.session_state[f'p_type_{i}']   = pt if pt in PTYPS else ""
        st.session_state[f'p_price_{i}']  = float(p.get('price',0))
        st.session_state[f'p_fanop_{i}']  = float(p.get('fanop',0))
        st.session_state[f'p_vafeas_{i}'] = float(p.get('vafeas',0))
        st.session_state[f'p_mix_{i}']    = float(p.get('mixanikos',0))
        st.session_state[f'p_il_{i}']     = float(p.get('ilgos',0))

    works_data = data.get('works',[])
    st.session_state['num_works'] = max(len(works_data), 1)
    WTYPS = ["","ΕΞΑΓΩΓΗ-ΤΟΠΟΘΕΤΗΣΗ","ΕΠΙΣΚΕΥΗ"]
    for i, w in enumerate(works_data):
        wt = w.get('type','')
        st.session_state[f'w_type_{i}']   = wt if wt in WTYPS else ""
        st.session_state[f'w_desc_{i}']   = w.get('desc','')
        st.session_state[f'w_fanop_{i}']  = float(w.get('fanop',0))
        st.session_state[f'w_vafeas_{i}'] = float(w.get('vafeas',0))
        st.session_state[f'w_mix_{i}']    = float(w.get('mixanikos',0))
        st.session_state[f'w_il_{i}']     = float(w.get('ilgos',0))

    st.session_state['_html_msg'] = (
        f"✅ Φορτώθηκαν: {len(parts_data)} ανταλλακτικά, {len(works_data)} εργασίες"
    )


def on_html_upload():
    f = st.session_state.get('html_upload')
    if f is None:
        return
    if st.session_state.get('_last_html') == f.name:
        return
    st.session_state['_last_html'] = f.name

    raw = f.read()
    html_content = None
    for enc in ('windows-1253','utf-8','latin-1'):
        try:
            html_content = raw.decode(enc)
            break
        except:
            continue
    if not html_content:
        st.session_state['_html_msg'] = "❌ Αδύνατη η ανάγνωση του αρχείου"
        return

    data, err = parse_html_report(html_content)
    if err:
        st.session_state['_html_msg'] = f"❌ {err}"
        return

    apply_to_session(data)


# ============================================================
# OCR ΑΔΕΙΑΣ ΚΥΚΛΟΦΟΡΙΑΣ ΜΕ GEMINI
# ============================================================
def ocr_adeia_kykloforias(img_bytes: bytes):
    """Στέλνει φωτογραφία άδειας στο Claude και επιστρέφει τα στοιχεία."""
    import os, json, re, base64
    try:
        api_key = st.secrets.get("ANTHROPIC_API_KEY","") or os.environ.get("ANTHROPIC_API_KEY","")
    except:
        api_key = os.environ.get("ANTHROPIC_API_KEY","")
    if not api_key:
        return None, "Δεν βρέθηκε ANTHROPIC_API_KEY στο secrets.toml"
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        # Ανίχνευση format από PIL
        import io as _io
        _img_check = PILImage.open(_io.BytesIO(img_bytes))
        _fmt = _img_check.format or "JPEG"
        media_type = "image/jpeg" if _fmt.upper() in ("JPEG","JPG") else "image/png"

        img_b64 = base64.standard_b64encode(img_bytes).decode("utf-8")

        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1000,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": img_b64,
                        },
                    },
                    {
                        "type": "text",
                        "text": """Αυτή είναι ελληνική άδεια κυκλοφορίας οχήματος.
Διάβασε προσεκτικά και επέστρεψε ΜΟΝΟ ένα JSON object.
Αν δεν μπορείς να διαβάσεις κάποιο πεδίο, βάλε κενό string "".
{
  "ar_kykloforias": "πινακίδα π.χ. ΑΤΗ8498",
  "ar_plaisiou": "αριθμός πλαισίου VIN",
  "idioktitis": "επώνυμο και όνομα ιδιοκτήτη",
  "marka": "μάρκα οχήματος",
  "montelo": "μοντέλο",
  "kyvika": "κυβισμός σε cc χωρίς μονάδα",
  "proti_adeia": "ημερομηνία πρώτης άδειας π.χ. 14/10/2011",
  "xrisi": "χρήση π.χ. ΕΙΧ ή ΦΙΧ"
}
Επέστρεψε ΜΟΝΟ το JSON, χωρίς άλλο κείμενο."""
                    }
                ],
            }]
        )
        text = msg.content[0].text.strip()
        text = re.sub(r"```json|```", "", text).strip()
        data = json.loads(text)
        return data, None
    except Exception as e:
        return None, str(e)


# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(page_title="Experts360 v2.1", page_icon="📋", layout="wide")

# ============================================================
# LOGIN SYSTEM
# ============================================================
import hashlib

# Χρήστες — αλλάξτε τους κωδικούς!
USERS = {
    "admin": hashlib.sha256("experts360".encode()).hexdigest(),
    "kostas": hashlib.sha256("gnomon2026".encode()).hexdigest(),
}

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def check_login(username: str, password: str) -> bool:
    return USERS.get(username) == hash_password(password)

def show_login():
    st.markdown("""
    <style>
    .login-box {
        max-width: 400px;
        margin: 80px auto;
        padding: 2rem;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
        background: white;
    }
    </style>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1,2,1])
    with col2:
        import os as _os_login
        if _os_login.path.exists("logo_experts360.png"):
            st.image("logo_experts360.png", width=200)
        st.markdown("### Experts360 v2.1")
        st.markdown("*Πραγματογνωμοσύνη*")
        st.markdown("---")

        username = st.text_input("👤 Όνομα χρήστη", key="login_user")
        password = st.text_input("🔒 Κωδικός", type="password", key="login_pass")

        if st.button("🔓 Είσοδος", type="primary", use_container_width=True):
            if check_login(username, password):
                st.session_state['logged_in'] = True
                st.session_state['username'] = username
                st.rerun()
            else:
                st.error("❌ Λάθος όνομα χρήστη ή κωδικός")

# Έλεγχος login
if not st.session_state.get('logged_in', False):
    show_login()
    st.stop()
import os as _os

# Εικαστικό ανάλογα με κατηγορία
_kat = st.session_state.get('kategoria', '🚗 Οχήματα')
if _kat == '🚗 Οχήματα':
    st.markdown("""<svg width="100%" viewBox="0 0 680 180" role="img" style="margin-bottom:8px;border-radius:12px">
<title>Εκθέσεις οχημάτων</title>
<rect x="0" y="0" width="680" height="180" fill="#e8f4fd" rx="12"/>
<rect x="0" y="135" width="680" height="45" fill="#c8d4dc"/>
<rect x="0" y="132" width="680" height="6" fill="#90a4ae"/>
<rect x="60" y="146" width="40" height="4" fill="white" opacity="0.7" rx="2"/>
<rect x="250" y="146" width="40" height="4" fill="white" opacity="0.7" rx="2"/>
<rect x="450" y="146" width="40" height="4" fill="white" opacity="0.7" rx="2"/>
<ellipse cx="80" cy="135" rx="110" ry="35" fill="#a5d6a7"/>
<ellipse cx="600" cy="137" rx="100" ry="30" fill="#81c784"/>
<rect x="25" y="100" width="5" height="30" fill="#795548"/>
<ellipse cx="27" cy="96" rx="14" ry="17" fill="#4caf50"/>
<rect x="620" y="105" width="5" height="28" fill="#795548"/>
<ellipse cx="622" cy="101" rx="12" ry="15" fill="#66bb6a"/>
<rect x="230" y="72" width="220" height="68" rx="10" fill="#1565c0"/>
<path d="M258 72 Q275 44 315 40 L375 40 Q412 44 422 72Z" fill="#1976d2"/>
<path d="M268 70 Q280 48 313 45 L368 45 Q398 50 410 70Z" fill="#b3e5fc" opacity="0.9"/>
<line x1="342" y1="45" x2="342" y2="140" stroke="#1251a3" stroke-width="1.5"/>
<circle cx="268" cy="138" r="20" fill="#212121"/>
<circle cx="268" cy="138" r="12" fill="#616161"/>
<circle cx="268" cy="138" r="5" fill="#9e9e9e"/>
<circle cx="412" cy="138" r="20" fill="#212121"/>
<circle cx="412" cy="138" r="12" fill="#616161"/>
<circle cx="412" cy="138" r="5" fill="#9e9e9e"/>
<rect x="447" y="87" width="11" height="7" rx="2" fill="#fff9c4"/>
<rect x="222" y="87" width="11" height="7" rx="2" fill="#ef9a9a"/>
<circle cx="595" cy="38" r="26" fill="#ffe082" opacity="0.9"/>
<line x1="595" y1="5" x2="595" y2="0" stroke="#ffca28" stroke-width="2.5" stroke-linecap="round"/>
<line x1="560" y1="38" x2="554" y2="38" stroke="#ffca28" stroke-width="2.5" stroke-linecap="round"/>
<line x1="630" y1="38" x2="636" y2="38" stroke="#ffca28" stroke-width="2.5" stroke-linecap="round"/>
<line x1="572" y1="15" x2="568" y2="10" stroke="#ffca28" stroke-width="2" stroke-linecap="round"/>
<line x1="618" y1="15" x2="622" y2="10" stroke="#ffca28" stroke-width="2" stroke-linecap="round"/>
</svg>""", unsafe_allow_html=True)
else:
    st.markdown("""<svg width="100%" viewBox="0 0 680 180" role="img" style="margin-bottom:8px;border-radius:12px">
<title>Εκθέσεις κτιρίων</title>
<rect x="0" y="0" width="680" height="180" fill="#f1f8e9" rx="12"/>
<rect x="0" y="140" width="680" height="40" fill="#c8d8c0"/>
<rect x="0" y="138" width="680" height="5" fill="#a5c49b"/>
<ellipse cx="80" cy="141" rx="60" ry="10" fill="#81c784" opacity="0.6"/>
<ellipse cx="600" cy="142" rx="70" ry="11" fill="#81c784" opacity="0.5"/>
<rect x="265" y="58" width="150" height="85" rx="4" fill="#eceff1" stroke="#b0bec5" stroke-width="1.5"/>
<line x1="265" y1="83" x2="415" y2="83" stroke="#cfd8dc" stroke-width="1"/>
<line x1="265" y1="108" x2="415" y2="108" stroke="#cfd8dc" stroke-width="1"/>
<rect x="278" y="65" width="20" height="14" rx="2" fill="#90caf9" opacity="0.8"/>
<rect x="308" y="65" width="20" height="14" rx="2" fill="#fff9c4" opacity="0.9"/>
<rect x="338" y="65" width="20" height="14" rx="2" fill="#90caf9" opacity="0.8"/>
<rect x="368" y="65" width="20" height="14" rx="2" fill="#fff9c4" opacity="0.9"/>
<rect x="278" y="89" width="20" height="14" rx="2" fill="#fff9c4" opacity="0.9"/>
<rect x="308" y="89" width="20" height="14" rx="2" fill="#90caf9" opacity="0.8"/>
<rect x="338" y="89" width="20" height="14" rx="2" fill="#fff9c4" opacity="0.9"/>
<rect x="368" y="89" width="20" height="14" rx="2" fill="#90caf9" opacity="0.8"/>
<rect x="304" y="116" width="72" height="27" rx="3" fill="#8d6e63"/>
<rect x="314" y="122" width="20" height="14" rx="1" fill="#a1887f" opacity="0.5"/>
<rect x="340" y="122" width="20" height="14" rx="1" fill="#a1887f" opacity="0.5"/>
<path d="M258 60 L340 26 L422 60Z" fill="#b0bec5" stroke="#90a4ae" stroke-width="1.5"/>
<rect x="385" y="32" width="11" height="26" rx="2" fill="#90a4ae"/>
<rect x="110" y="90" width="75" height="55" rx="3" fill="#f5f5f5" stroke="#cfd8dc" stroke-width="1"/>
<line x1="110" y1="108" x2="185" y2="108" stroke="#e0e0e0" stroke-width="1"/>
<rect x="120" y="95" width="14" height="11" rx="2" fill="#90caf9" opacity="0.7"/>
<rect x="143" y="95" width="14" height="11" rx="2" fill="#fff9c4" opacity="0.8"/>
<rect x="120" y="113" width="14" height="11" rx="2" fill="#fff9c4" opacity="0.8"/>
<rect x="143" y="113" width="14" height="11" rx="2" fill="#90caf9" opacity="0.7"/>
<rect x="130" y="128" width="22" height="18" rx="2" fill="#795548"/>
<rect x="105" y="84" width="85" height="9" rx="2" fill="#b0bec5"/>
<rect x="498" y="85" width="90" height="58" rx="3" fill="#f5f5f5" stroke="#cfd8dc" stroke-width="1"/>
<line x1="498" y1="104" x2="588" y2="104" stroke="#e0e0e0" stroke-width="1"/>
<line x1="498" y1="123" x2="588" y2="123" stroke="#e0e0e0" stroke-width="1"/>
<rect x="508" y="90" width="15" height="12" rx="2" fill="#fff9c4" opacity="0.9"/>
<rect x="532" y="90" width="15" height="12" rx="2" fill="#90caf9" opacity="0.8"/>
<rect x="558" y="90" width="15" height="12" rx="2" fill="#fff9c4" opacity="0.9"/>
<rect x="508" y="108" width="15" height="12" rx="2" fill="#90caf9" opacity="0.8"/>
<rect x="532" y="108" width="15" height="12" rx="2" fill="#fff9c4" opacity="0.9"/>
<rect x="558" y="108" width="15" height="12" rx="2" fill="#90caf9" opacity="0.8"/>
<rect x="520" y="128" width="30" height="16" rx="2" fill="#795548"/>
<rect x="492" y="79" width="102" height="9" rx="2" fill="#b0bec5"/>
<ellipse cx="160" cy="35" rx="38" ry="16" fill="white" opacity="0.8"/>
<ellipse cx="188" cy="28" rx="28" ry="18" fill="white" opacity="0.8"/>
<ellipse cx="205" cy="40" rx="23" ry="13" fill="white" opacity="0.8"/>
</svg>""", unsafe_allow_html=True)

# Αρχικοποίηση βάσης
if 'db_ready' not in st.session_state:
    ok, db_type = init_db()
    st.session_state['db_ready'] = ok
    st.session_state['db_type'] = db_type if ok else "error"


db_icon = "🟢" if st.session_state.get('db_type') in ('sqlite','postgres','mysql') else "🔴"
db_label = {"sqlite":"SQLite (τοπικά)", "postgres":"PostgreSQL (cloud)", "mysql":"MySQL (experts360.gr)"}.get(
    st.session_state.get('db_type',''), "Σφάλμα")


# Debug: εμφάνιση URL status
try:
    import streamlit as _st2
    _url = _st2.secrets.get("GNOMON_DB_URL","")
    if _url:
        st.session_state['_db_url_found'] = True
except:
    _url = ""

st.caption(f"{db_icon} Βάση: {db_label}")

st.markdown("---")

# ============================================================
# SESSION STATE DEFAULTS
# ============================================================
for key, val in [('num_parts',3),('num_works',2),('num_visits',1),
                 ('_html_msg',''),('_last_html',''),
                 ('current_ekthesi_id', None),('db_page','form')]:
    if key not in st.session_state:
        st.session_state[key] = val

# ============================================================
# HTML IMPORT
# ============================================================
# ============================================================
# SIDEBAR - ΠΛΟΗΓΗΣΗ & ΑΝΑΖΗΤΗΣΗ
# ============================================================
with st.sidebar:
    import os as _os2
    if _os2.path.exists("logo_experts360.png"):
        st.image("logo_experts360.png", width=180)
    else:
        st.markdown("### Experts360")
    st.markdown("---")

    # Κατηγορία
    kategoria = st.radio("🗂️ Κατηγορία", ["🚗 Οχήματα", "🏢 Κτίρια"], key="kategoria")
    st.markdown("---")

    PAGES = ["📝 Νέα Έκθεση", "🔍 Αναζήτηση", "📊 Στατιστικά", "⚙️ Ρυθμίσεις"]
    if 'nav_index' not in st.session_state:
        st.session_state['nav_index'] = 0
    page = st.radio("📌 Μενού", PAGES, index=st.session_state['nav_index'])
    st.session_state['nav_index'] = PAGES.index(page)

    st.markdown("---")

    if st.session_state.get('current_ekthesi_id'):
        st.info(f"📂 Ανοιχτή: **#{st.session_state['current_ekthesi_id']}**")
        if st.button("🆕 Νέα (καθαρή)", use_container_width=True):
            for k in list(st.session_state.keys()):
                if k not in ('db_ready','db_type','nav_index',
                             'num_parts','num_works','num_visits'):
                    del st.session_state[k]
            st.session_state['current_ekthesi_id'] = None
            st.session_state['num_parts'] = 3
            st.session_state['num_works'] = 2
            st.session_state['num_visits'] = 1
            st.rerun()

    st.markdown("---")
    st.caption("Experts360 v2.1")
    st.markdown("---")
    st.caption(f"👤 {st.session_state.get('username','')}")
    if st.button("🚪 Αποσύνδεση", use_container_width=True):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

# Routing για Κτίρια
if st.session_state.get('kategoria') == "🏢 Κτίρια":
    from ktirion_ui import show_ktirion_tab
    show_ktirion_tab(page)
    st.stop()

# Ρυθμίσεις - Διαχείριση μαρκών/μοντέλων
if page == "⚙️ Ρυθμίσεις":
    st.subheader("⚙️ Ρυθμίσεις — Μάρκες & Μοντέλα")

    tab1, tab2 = st.tabs(["➕ Προσθήκη", "🗑️ Διαγραφή"])

    with tab1:
        st.markdown("#### Προσθήκη Νέας Μάρκας")
        c1, c2 = st.columns([2,1])
        with c1:
            new_marka_input = st.text_input("Νέα Μάρκα", placeholder="π.χ. BYD", key="new_marka_input")
        with c2:
            st.write("")
            st.write("")
            if st.button("➕ Προσθήκη Μάρκας", use_container_width=True):
                if new_marka_input.strip():
                    ok, err = add_custom_marka(new_marka_input.strip())
                    if ok: st.success(f"✅ Προστέθηκε: {new_marka_input}")
                    else: st.error(f"❌ {err}")
                else:
                    st.warning("Γράψε μάρκα")

        st.markdown("#### Προσθήκη Μοντέλου σε Μάρκα")
        _all_m = sorted(set(get_markes() + get_custom_markes()))
        c3, c4, c5 = st.columns([2,2,1])
        with c3:
            sel_marka_mod = st.selectbox("Μάρκα", [""] + _all_m, key="sel_marka_mod")
        with c4:
            new_montelo_input = st.text_input("Νέο Μοντέλο", placeholder="π.χ. Atto 3", key="new_montelo_input")
        with c5:
            st.write("")
            st.write("")
            if st.button("➕ Μοντέλο", use_container_width=True):
                if sel_marka_mod and new_montelo_input.strip():
                    ok, err = add_custom_montelo(sel_marka_mod, new_montelo_input.strip())
                    if ok: st.success(f"✅ {sel_marka_mod} → {new_montelo_input}")
                    else: st.error(f"❌ {err}")
                else:
                    st.warning("Επίλεξε μάρκα και γράψε μοντέλο")

    with tab2:
        st.markdown("#### Custom Μάρκες & Μοντέλα")
        custom_m = get_custom_markes()
        if not custom_m:
            st.info("Δεν έχεις προσθέσει custom μάρκες ακόμα.")
        else:
            for cm in custom_m:
                with st.expander(f"🚗 {cm}"):
                    if st.button(f"🗑️ Διαγραφή μάρκας {cm}", key=f"del_marka_{cm}"):
                        delete_custom_vehicle(cm)
                        st.rerun()
                    montela_cm = get_custom_montela(cm)
                    if montela_cm:
                        st.markdown("**Μοντέλα:**")
                        for mo in montela_cm:
                            mc1, mc2 = st.columns([4,1])
                            mc1.write(f"• {mo}")
                            if mc2.button("🗑️", key=f"del_mo_{cm}_{mo}"):
                                delete_custom_vehicle(cm, mo)
                                st.rerun()
    st.stop()

# Αν επιλέξει Αναζήτηση
if page == "🔍 Αναζήτηση":
    st.subheader("🔍 Αναζήτηση Εκθέσεων")

    col1, col2 = st.columns([3,1])
    with col1:
        q = st.text_input("Αναζήτηση (αρ.ζημίας, ιδιοκτήτης, πινακίδα, μάρκα...)",
                          placeholder="π.χ. Toyota ή ΓΙΔΟΠΟΥΛΟΣ")
    with col2:
        status_filter = st.selectbox("Status", ["","draft","final","archived"],
                                     format_func=lambda x: {"":"Όλα","draft":"Προσχέδιο",
                                     "final":"Τελική","archived":"Αρχείο"}.get(x,x))

    results = search_ektheseis(query=q, status=status_filter, limit=100)

    st.write(f"**{len(results)} εκθέσεις**")

    for r in results:
        status_color = {"draft":"🟡","final":"🟢","archived":"⚫"}.get(r.get('status',''),'⚪')
        with st.expander(
            f"{status_color} #{r['id']} | {r.get('ar_zimias','')} | "
            f"{r.get('idioktitis','')} | {r.get('marka','')} {r.get('montelo','')} "
            f"| {r.get('hm_entolhs','')}"
        ):
            col1, col2, col3 = st.columns(3)
            with col1:
                if st.button("📂 Φόρτωση", key=f"load_{r['id']}", use_container_width=True):
                    ekthesi = load_ekthesi(r['id'])
                    if ekthesi:
                        st.session_state['current_ekthesi_id'] = r['id']
                        st.session_state['ar_zimias']        = ekthesi.get('ar_zimias','')
                        st.session_state['topos_atyxhmatos'] = ekthesi.get('topos_atyx','')
                        st.session_state['idioktitis']       = ekthesi.get('idioktitis','')
                        st.session_state['ar_kykloforias']   = ekthesi.get('ar_kykl','')
                        st.session_state['marka']            = ekthesi.get('marka','')
                        st.session_state['montelo']          = ekthesi.get('montelo','')
                        st.session_state['kyvika']           = ekthesi.get('kyvika','')
                        st.session_state['xrisi']            = ekthesi.get('xrisi','')
                        st.session_state['proti_adeia']      = ekthesi.get('proti_adeia','')
                        st.session_state['xiliometrites']    = ekthesi.get('xiliom','')
                        st.session_state['axia']             = int(ekthesi.get('axia') or 0)
                        st.session_state['vd_text_0']        = ekthesi.get('visit_date','')
                        st.session_state['visit_place_0']    = ekthesi.get('visit_place','')

                        parts_db = ekthesi.get('parts', [])
                        st.session_state['num_parts'] = max(len(parts_db), 3)
                        for i, p in enumerate(parts_db):
                            st.session_state[f'p_name_{i}']   = p.get('name','')
                            st.session_state[f'p_type_{i}']   = p.get('type','')
                            st.session_state[f'p_price_{i}']  = float(p.get('price',0))
                            st.session_state[f'p_fanop_{i}']  = float(p.get('fanop',0))
                            st.session_state[f'p_vafeas_{i}'] = float(p.get('vafeas',0))
                            st.session_state[f'p_mix_{i}']    = float(p.get('mixanikos',0))
                            st.session_state[f'p_il_{i}']     = float(p.get('ilgos',0))

                        works_db = ekthesi.get('works', [])
                        st.session_state['num_works'] = max(len(works_db), 2)
                        for i, w in enumerate(works_db):
                            st.session_state[f'w_type_{i}']   = w.get('type','')
                            st.session_state[f'w_desc_{i}']   = w.get('desc','')
                            st.session_state[f'w_fanop_{i}']  = float(w.get('fanop',0))
                            st.session_state[f'w_vafeas_{i}'] = float(w.get('vafeas',0))
                            st.session_state[f'w_mix_{i}']    = float(w.get('mixanikos',0))
                            st.session_state[f'w_il_{i}']     = float(w.get('ilgos',0))

                        # Πηγαίνουμε στη φόρμα
                        st.session_state['nav_index'] = 0
                        st.rerun()

            with col2:
                new_s = "final" if r.get('status') == "draft" else "draft"
                lbl = "✅ Οριστικοποίηση" if new_s == "final" else "🔄 Επαναφορά"
                if st.button(lbl, key=f"status_{r['id']}", use_container_width=True):
                    update_status(r['id'], new_s)
                    st.rerun()

            with col3:
                if st.button("🗑️ Διαγραφή", key=f"del_{r['id']}", use_container_width=True,
                             type="secondary"):
                    ok, err = delete_ekthesi(r['id'])
                    if ok:
                        st.success("Διαγράφηκε")
                        st.rerun()
                    else:
                        st.error(err)

            # Ιστορικό
            hist = get_history(r['id'])
            if hist:
                st.caption("**Ιστορικό:**")
                for h in hist[:5]:
                    st.caption(f"  {h.get('created_at','')} — {h.get('action','')} {h.get('details','')[:60]}")

    st.stop()

# Αν επιλέξει Στατιστικά
if page == "📊 Στατιστικά":
    st.subheader("📊 Στατιστικά")
    stats = get_statistics()

    col1, col2, col3 = st.columns(3)
    col1.metric("Σύνολο εκθέσεων", stats.get('total', 0))
    col2.metric("Τελευταίες 30 μέρες", stats.get('last_30_days', 0))
    col3.metric("Συνολική αξία", f"{stats.get('total_axia', 0):,.0f} €")

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Top 5 Μάρκες:**")
        for m in stats.get('top_markes', []):
            st.write(f"- {m.get('marka','')}: **{m.get('cnt',0)}** εκθέσεις")

    with col2:
        st.markdown("**Ανά Status:**")
        labels = {"draft":"Προσχέδιο","final":"Τελική","archived":"Αρχείο"}
        for s, cnt in stats.get('by_status', {}).items():
            st.write(f"- {labels.get(s,s)}: **{cnt}**")

    st.markdown("---")
    if st.button("⬇️ Export όλων σε JSON (backup)"):
        j = export_all_json()
        st.download_button("📥 Κατέβασμα JSON", data=j,
                           file_name=f"gnomon_backup_{datetime.now().strftime('%Y%m%d')}.json",
                           mime="application/json")

    st.stop()

# ============================================================
# ΕΠΙΛΟΓΗ ΑΣΦΑΛΙΣΤΙΚΗΣ
# ============================================================
col_asf, _ = st.columns([2, 4])
with col_asf:
    asfalistiki = st.selectbox(
        "🏢 Ασφαλιστική Εταιρεία",
        options=["INTERLIFE", "ΕΘΝΙΚΗ ΑΣΦΑΛΙΣΤΙΚΗ", "APEIRON ΑΣΦΑΛΙΣΤΙΚΗ"],
        key="asfalistiki"
    )
st.markdown("---")

st.subheader("📥 Εισαγωγή από HTML Έκθεση")
st.caption("Ανεβάστε το HTML αρχείο από το σύστημα GNOMON για αυτόματη συμπλήρωση.")

st.file_uploader(
    "Επιλέξτε αρχείο HTML έκθεσης",
    type=['html','htm'],
    key="html_upload",
    on_change=on_html_upload
)

if st.session_state['_html_msg']:
    if st.session_state['_html_msg'].startswith('✅'):
        st.success(st.session_state['_html_msg'])
    else:
        st.error(st.session_state['_html_msg'])

st.markdown("---")

# ============================================================
# ΣΤΟΙΧΕΙΑ ΑΤΥΧΗΜΑΤΟΣ
# ============================================================
st.subheader("🚗 Στοιχεία Ατυχήματος")
col1, col2 = st.columns(2)
with col1:
    ar_zimias        = st.text_input("Αρ. Ζημίας", placeholder="π.χ. 37150/25", key="ar_zimias")
    hm_atyxhmatos    = st.text_input("Ημ/νία Ατυχήματος", placeholder="ΗΗ/ΜΜ/ΕΕΕΕ", key="hm_atyxhmatos")
with col2:
    hm_entolhs       = st.text_input("Ημ/να Εντολής", placeholder="ΗΗ/ΜΜ/ΕΕΕΕ", key="hm_entolhs")
    topos_atyxhmatos = st.text_input("Τόπος Ατυχήματος", key="topos_atyxhmatos")
st.markdown("---")

# ============================================================
# ΣΤΟΙΧΕΙΑ ΟΧΗΜΑΤΟΣ
# ============================================================
# Έλεγχος αν υπάρχει Gemini key
try:
    _has_gemini = bool(st.secrets.get("ANTHROPIC_API_KEY","") or st.secrets.get("GEMINI_API_KEY",""))
except:
    _has_gemini = False

st.subheader("🚙 Στοιχεία Επιθεωρούμενου Οχήματος")

# === OCR ΑΔΕΙΑΣ ΚΥΚΛΟΦΟΡΙΑΣ ===
if _has_gemini:
    with st.expander("📷 Αυτόματη συμπλήρωση από φωτογραφία άδειας κυκλοφορίας", expanded=False):
        adeia_files = st.file_uploader(
            "Ανεβάστε φωτογραφίες άδειας κυκλοφορίας (1-3 σελίδες)",
            type=['jpg','jpeg','png'],
            accept_multiple_files=True,
            key="adeia_upload"
        )
        if adeia_files:
            prev_cols = st.columns(len(adeia_files))
            for idx, f in enumerate(adeia_files):
                with prev_cols[idx]:
                    st.image(f, caption=f"Σελίδα {idx+1}", use_container_width=True)

            if st.button("🤖 Ανάγνωση με AI", type="primary", use_container_width=True):
                with st.spinner(f"Αναγνώριση {len(adeia_files)} σελίδων..."):
                    merged = {}
                    for idx2, f in enumerate(adeia_files):
                        f.seek(0)
                        data_ocr, err_ocr = ocr_adeia_kykloforias(f.read())
                        if err_ocr:
                            st.warning(f"Σελίδα {idx2+1}: {err_ocr}")
                            continue
                        if data_ocr:
                            for k, v in data_ocr.items():
                                if v and not merged.get(k):
                                    merged[k] = v

                if merged:
                    # Χρησιμοποιούμε pending pattern για να αποφύγουμε
                    # το "cannot modify after widget instantiated" error
                    for field in ['idioktitis','ar_kykloforias','ar_plaisiou',
                                  'proti_adeia','kyvika','xrisi']:
                        if merged.get(field):
                            st.session_state[f'_ocr_pending_{field}'] = merged[field]

                    from vehicles import get_markes, get_montela
                    raw_m = merged.get('marka','')
                    markes_l = get_markes()
                    matched_m = next((m for m in markes_l if m.lower()==raw_m.lower()), raw_m)
                    st.session_state['_ocr_pending_marka'] = matched_m

                    raw_mo = merged.get('montelo','')
                    montela_l = get_montela(matched_m)
                    matched_mo = next((m for m in montela_l if m.lower()==raw_mo.lower()), raw_mo)
                    st.session_state['_ocr_pending_montelo'] = matched_mo

                    st.session_state['_ocr_show_results'] = merged
                    st.rerun()

                if st.session_state.get('_ocr_show_results'):
                    merged_show = st.session_state.pop('_ocr_show_results')
                    with st.expander("📋 Στοιχεία που βρέθηκαν", expanded=True):
                        labels = {
                            'idioktitis':'Ιδιοκτήτης','ar_kykloforias':'Πινακίδα',
                            'ar_plaisiou':'Αρ. Πλαισίου (VIN)','marka':'Μάρκα',
                            'montelo':'Μοντέλο','kyvika':'Κυβικά',
                            'proti_adeia':'1η Άδεια','xrisi':'Χρήση'
                        }
                        for k, v in merged_show.items():
                            if v:
                                st.write(f"**{labels.get(k,k)}:** {v}")
                    st.success("✅ Συμπληρώθηκαν τα στοιχεία!")
else:
    st.info("💡 Προσθέστε ANTHROPIC_API_KEY στο secrets.toml για αυτόματη ανάγνωση άδειας.")

# Εφαρμογή VIN pending values πριν τα widgets
if st.session_state.get('_vin_pending_marka'):
    st.session_state['marka']   = st.session_state.pop('_vin_pending_marka')
if st.session_state.get('_vin_pending_montelo'):
    st.session_state['montelo'] = st.session_state.pop('_vin_pending_montelo')
if st.session_state.get('_vin_pending_kyvika'):
    st.session_state['kyvika']  = st.session_state.pop('_vin_pending_kyvika') + ' cc'
if st.session_state.get('_vin_pending_xrisi'):
    st.session_state['xrisi']   = st.session_state.pop('_vin_pending_xrisi')

# Εφαρμογή OCR pending values πριν τα widgets
for _ocr_field in ['idioktitis','ar_kykloforias','ar_plaisiou','proti_adeia','kyvika','xrisi']:
    _ocr_key = f'_ocr_pending_{_ocr_field}'
    if st.session_state.get(_ocr_key):
        st.session_state[_ocr_field] = st.session_state.pop(_ocr_key)
if st.session_state.get('_ocr_pending_marka'):
    st.session_state['marka']   = st.session_state.pop('_ocr_pending_marka')
if st.session_state.get('_ocr_pending_montelo'):
    st.session_state['montelo'] = st.session_state.pop('_ocr_pending_montelo')
idioktitis = st.text_input("Ιδιοκτήτης", placeholder="Επώνυμο Όνομα", key="idioktitis")
col1, col2, col3 = st.columns(3)
with col1:
    ar_kykloforias = st.text_input("Αρ. Κυκλοφορίας", placeholder="ΙΡΕ-5840", key="ar_kykloforias")
    ar_plaisiou_raw = st.text_input("Αρ. Πλαισίου (VIN)", placeholder="π.χ. VSSZZZ6KZ1R049865",
                                    max_chars=17, key="ar_plaisiou")
    ar_plaisiou = ar_plaisiou_raw.strip().upper()
    if ar_plaisiou_raw and len(ar_plaisiou) < 17:
        st.caption(f"⚠️ VIN: {len(ar_plaisiou)}/17 χαρακτήρες")
    elif ar_plaisiou and len(ar_plaisiou) == 17:
        st.caption("✅ VIN έτοιμο για αναζήτηση")
    _all_markes = sorted(set(get_markes() + get_custom_markes()))
    _markes_list = [""] + _all_markes
    marka = st.selectbox("Μάρκα", options=_markes_list,
                         index=_markes_list.index(st.session_state.get('marka','')) if st.session_state.get('marka','') in _markes_list else 0,
                         key="marka")
    proti_adeia    = st.text_input("1η Άδεια Κυκλοφορίας", placeholder="π.χ. 02/05/1965", key="proti_adeia")
    xiliometrites  = st.text_input("Ένδειξη Χιλιομετρητή", key="xiliometrites")
with col2:
    xrisi   = st.text_input("Χρήση", placeholder="ΕΙΧ", key="xrisi")
    kyvika  = st.text_input("Κυβικά", key="kyvika")
    hm_kteo = st.text_input("Ημ/νία ΚΤΕΟ", placeholder="ΗΗ/ΜΜ/ΕΕΕΕ", key="hm_kteo")
    axia    = st.number_input("Αξία (€)", min_value=0, step=100, key="axia")
with col3:
    _all_montela = sorted(set(get_montela(marka) + get_custom_montela(marka))) if marka else []
    _montela_list = [""] + _all_montela
    montelo = st.selectbox("Μοντέλο", options=_montela_list,
                           index=_montela_list.index(st.session_state.get('montelo','')) if st.session_state.get('montelo','') in _montela_list else 0,
                           key="montelo")

# === ΕΠΙΠΛΕΟΝ ΠΕΔΙΑ (μόνο ΕΘΝΙΚΗ/APEIRON) ===
_asfalistiki_sel = st.session_state.get('asfalistiki','INTERLIFE')
if _asfalistiki_sel in ('ΕΘΝΙΚΗ ΑΣΦΑΛΙΣΤΙΚΗ','APEIRON ΑΣΦΑΛΙΣΤΙΚΗ'):
    st.markdown("---")
    st.markdown("#### 🔎 Επιπλέον Στοιχεία Οχήματος")
    ea1, ea2, ea3 = st.columns(3)
    with ea1:
        xroma       = st.text_input("Χρώμα", key="xroma")
        kaysimo     = st.text_input("Καύσιμο", placeholder="Βενζίνη/Diesel/Hybrid", key="kaysimo")
        katast_oxima = st.text_input("Κατάσταση Οχήματος", key="katast_oxima")
    with ea2:
        ixni_xromatos  = st.text_input("Ίχνη Χρωμάτων", key="ixni_xromatos")
        fora_atyxima   = st.text_input("Φορά Ατυχήματος", key="fora_atyxima")
        elastikon_simeio = st.text_input("Σημείο Ελαστικών", key="elastikon_simeio")
    with ea3:
        tilefono       = st.text_input("Τηλέφωνο Ιδιοκτήτη", key="tilefono")
        _kind_sel = st.selectbox("Κίνδυνος", ["","ΣΑΠ","ΤΡΙΤΟΣ","ΚΑΛΥΨΗ"], key="kindynos_sel")
        kindynos = st.text_input("ή γράψε κίνδυνο", value=_kind_sel, key="kindynos")

    st.markdown("#### 🔧 Στοιχεία Συνεργείου")

    # Autocomplete συνεργείου
    _syn_list = get_synergeio_eponimies()
    if _syn_list:
        _syn_options = ["-- Νέο συνεργείο --"] + _syn_list
        _syn_sel = st.selectbox("📋 Επιλογή από προηγούμενα συνεργεία",
                                _syn_options, key="syn_select",
                                index=0)
        if _syn_sel != "-- Νέο συνεργείο --":
            _syn_data = get_synergeio_full(_syn_sel)
            if _syn_data and st.button("📥 Φόρτωση στοιχείων", key="syn_load"):
                for f, v in [('synergeio_eponimia', _syn_data.get('eponimia','')),
                             ('synergeio_dieuthinsi', _syn_data.get('dieuthinsi','')),
                             ('synergeio_tilefono', _syn_data.get('tilefono','')),
                             ('synergeio_kinito', _syn_data.get('kinito','')),
                             ('synergeio_fax', _syn_data.get('fax','')),
                             ('synergeio_mail', _syn_data.get('mail',''))]:
                    st.session_state[f] = v or ''
                st.rerun()

    sb1, sb2, sb3 = st.columns(3)
    with sb1:
        synergeio_eponimia   = st.text_input("Επωνυμία Συνεργείου", key="synergeio_eponimia")
        synergeio_dieuthinsi = st.text_input("Διεύθυνση", key="synergeio_dieuthinsi")
    with sb2:
        synergeio_tilefono = st.text_input("Τηλέφωνο Συνεργείου", key="synergeio_tilefono")
        synergeio_kinito   = st.text_input("Κινητό Συνεργείου", key="synergeio_kinito")
    with sb3:
        synergeio_fax  = st.text_input("Fax", key="synergeio_fax")
        synergeio_mail = st.text_input("Email Συνεργείου", key="synergeio_mail")
else:
    xroma = kaysimo = katast_oxima = ixni_xromatos = fora_atyxima = ''
    elastikon_simeio = tilefono = kindynos = ''
    synergeio_eponimia = synergeio_dieuthinsi = synergeio_tilefono = ''
    synergeio_kinito = synergeio_fax = synergeio_mail = ''

# === VIN DECODER ===
_vin_val = ar_plaisiou.strip() if ar_plaisiou else ""
if len(_vin_val) >= 17:
    if st.button("🔎 Αναζήτηση VIN", key="vin_search"):
        with st.spinner("Αναζήτηση στοιχείων VIN..."):
            try:
                import requests as req_lib
                import hashlib

                # Διαβάζουμε API key
                try:
                    vd_key = st.secrets.get("VINDECODER_API_KEY","")
                except:
                    vd_key = ""

                vin_results = {}

                if vd_key:
                    # === vindecoder.eu (κύρια πηγή για ευρωπαϊκά) ===
                    id_str = f"{_vin_val}|{vd_key}|decode"
                    api_id = hashlib.sha1(id_str.encode()).hexdigest()[:9]
                    vd_url = f"https://api.vindecoder.eu/3.2/{vd_key}/{api_id}/decode/{_vin_val}.json"
                    resp = req_lib.get(vd_url, timeout=15)
                    if resp.status_code == 200:
                        vd_data = resp.json()
                        decode = vd_data.get('decode', [])
                        for item in decode:
                            label = item.get('label','')
                            val   = item.get('value','')
                            if val and val not in ('--','null','None',''):
                                vin_results[label] = str(val)

                if not vin_results:
                    # === Fallback: NHTSA ===
                    nhtsa_url = f"https://vpic.nhtsa.dot.gov/api/vehicles/decodevin/{_vin_val}?format=json"
                    resp2 = req_lib.get(nhtsa_url, timeout=15)
                    if resp2.status_code == 200:
                        nhtsa_data = resp2.json()
                        for x in nhtsa_data['Results']:
                            if x['Value'] and x['Value'] not in ('Not Applicable','null',None,''):
                                vin_results[x['Variable']] = x['Value']

                if vin_results:
                    st.session_state['_vin_results'] = vin_results

                    # Mapping vindecoder -> pending (υποστηρίζει και NHTSA format)
                    def get_val_multi(d, *keys):
                        for k in keys:
                            if d.get(k): return d[k]
                        return ''

                    st.session_state['_vin_pending_marka']   = get_val_multi(vin_results, 'Make', 'make').title()
                    st.session_state['_vin_pending_montelo'] = get_val_multi(vin_results, 'Model', 'model').title()
                    cc = get_val_multi(vin_results, 'Engine Displacement (ccm)', 'Displacement (CC)', 'engine_cc')
                    st.session_state['_vin_pending_kyvika']  = str(int(float(cc))) if cc else ''
                    st.session_state['_vin_pending_xrisi']   = get_val_multi(vin_results, 'Body', 'Body Class', 'body')
                    st.rerun()
                else:
                    st.error("❌ Δεν βρέθηκαν στοιχεία για αυτό το VIN")

            except Exception as timeout_e:
                if 'timeout' in str(timeout_e).lower() or 'timed out' in str(timeout_e).lower():
                    st.error("⏱️ Timeout — δοκίμασε ξανά σε λίγο.")
                else:
                    raise timeout_e
            except Exception as e:
                st.error(f"Σφάλμα: {e}")

if st.session_state.get('_vin_results'):
    vr = st.session_state['_vin_results']
    with st.expander("📋 Στοιχεία VIN — επίλεξε για παρατηρήσεις", expanded=True):
        vin_labels = {
            # vindecoder.eu labels
            'Make': 'Μάρκα', 'Model': 'Μοντέλο', 'Model Year': 'Έτος',
            'Body': 'Αμάξωμα', 'Engine Displacement (ccm)': 'Κυβισμός (cc)',
            'Fuel Type': 'Καύσιμο', 'Number of Cylinders': 'Κύλινδροι',
            'Transmission': 'Κιβώτιο', 'Drive': 'Κίνηση',
            'Country': 'Χώρα κατασκευής', 'Plant': 'Εργοστάσιο',
            # NHTSA labels (fallback)
            'Body Class': 'Αμάξωμα', 'Displacement (CC)': 'Κυβισμός (cc)',
            'Fuel Type - Primary': 'Καύσιμο', 'Manufacturer Name': 'Κατασκευαστής',
            'Plant Country': 'Χώρα κατασκευής', 'Model Year': 'Έτος κατασκευής',
        }
        selected_vin = []
        vc1, vc2 = st.columns(2)
        for vidx, (eng_key, gr_label) in enumerate(vin_labels.items()):
            if eng_key in vr:
                with (vc1 if vidx%2==0 else vc2):
                    if st.checkbox(f"**{gr_label}:** {vr[eng_key]}", key=f"vin_chk_{eng_key}"):
                        selected_vin.append(f"{gr_label}: {vr[eng_key]}")
        if selected_vin:
            if st.button("📝 Προσθήκη στις Παρατηρήσεις", key="vin_to_par"):
                st.session_state['_vin_par_text'] = "Στοιχεία VIN: " + " | ".join(selected_vin)
                st.rerun()

# === ΕΜΠΟΡΙΚΗ ΑΞΙΑ - CAR.GR + ΙΣΤΟΡΙΚΟ ===
if marka and montelo:
    import urllib.parse
    car_q   = urllib.parse.quote(f"{marka} {montelo}")
    car_url = f"https://www.car.gr/classifieds/cars/?lang=el&q={car_q}"

    axia_stats = get_axia_stats(marka, montelo)

    with st.expander(f"🔍 Εμπορική αξία {marka} {montelo}", expanded=True):
        col_a, col_b = st.columns([1,1])
        with col_a:
            st.markdown(f"**📱 Αναζήτηση αγγελιών:**")
            st.markdown(f"[▶ Άνοιγμα στο car.gr]({car_url}){{target='_blank'}}")
            st.caption("Κάνε screenshot και επισύναψέ το παρακάτω στις φωτογραφίες")
        with col_b:
            if axia_stats:
                st.markdown(f"**📊 Από τις δικές σου εκθέσεις ({axia_stats['count']} εκθέσεις):**")
                st.markdown(f"- Μέση αξία: **{axia_stats['mean']:,} €**")
                st.markdown(f"- Εύρος: {axia_stats['min']:,} — {axia_stats['max']:,} €")
                st.markdown(f"- Τελευταία: {axia_stats['last']:,} € ({axia_stats.get('last_date','')})")
            else:
                st.info("Δεν υπάρχουν ακόμα εκθέσεις για αυτό το όχημα.")

st.markdown("---")

# ============================================================
# ΕΠΙΣΚΕΨΕΙΣ
# ============================================================
st.subheader("🔍 Στοιχεία Επιθεώρησης")
c1, c2, _ = st.columns([1,1,4])
with c1:
    if st.button("➕ Προσθήκη επίσκεψης", use_container_width=True):
        st.session_state.num_visits += 1; st.rerun()
with c2:
    if st.button("➖ Αφαίρεση επίσκεψης", use_container_width=True,
                 disabled=st.session_state.num_visits <= 1):
        st.session_state.num_visits -= 1; st.rerun()

visits = []
for i in range(st.session_state.num_visits):
    c1, c2 = st.columns([2,3])
    with c1:
        # Πάντα text_input για ευκολία (δέχεται και παλιές ημερομηνίες)
        vd = st.text_input(f"Ημ/νία {i+1}ης Επίσκεψης", placeholder="π.χ. 08/05/2026",
                           key=f"vd_text_{i}")
    with c2:
        vp = st.text_input(
            f"Τόπος {i+1}ης Επίσκεψης" if i==0 else f"Τόπος {i+1}ης Επίσκεψης (προαιρετικό)",
            key=f"visit_place_{i}"
        )
    visits.append({"date_text": vd, "place": vp})
st.markdown("---")

# ============================================================
# ΑΝΤΑΛΛΑΚΤΙΚΑ
# ============================================================
st.subheader("🔧 Ανταλλακτικά")
st.caption("Η λέξη 'ΑΝΤΙΚΑΤΑΣΤΑΣΗ' μπαίνει αυτόματα.")

parts = []
hc = st.columns([3,1,1,1,1,1,1])
for col, lbl in zip(hc, ["**Ανταλλακτικό**","**Τύπος**","**Τιμή**","**Φανοπ.**","**Βαφέας**","**Μηχ/κός**","**Ηλ/γος**"]):
    col.markdown(lbl)

for i in range(st.session_state.num_parts):
    cols = st.columns([3,1,1,1,1,1,1])
    with cols[0]:
        name = st.text_input("Ανταλλακτικό", key=f"p_name_{i}",
                             label_visibility="collapsed", placeholder="π.χ. ΠΡΟΦΥΛΑΚΤΗΡΑΣ ΠΙΣΩ")
    with cols[1]:
        ptype = st.text_input("ΓΝ/ΙΜ/ΜΤΧ",
                              key=f"p_type_{i}", label_visibility="collapsed",
                              placeholder="ΓΝ")
    with cols[2]:
        price = st.number_input("Τιμή", min_value=0.0, step=0.01,
                                key=f"p_price_{i}", label_visibility="collapsed")
    with cols[3]:
        fanop = st.number_input("Φανοπ.", min_value=0.0, step=0.01,
                                key=f"p_fanop_{i}", label_visibility="collapsed")
    with cols[4]:
        vafeas = st.number_input("Βαφέας", min_value=0.0, step=0.01,
                                 key=f"p_vafeas_{i}", label_visibility="collapsed")
    with cols[5]:
        mix = st.number_input("Μηχ/κός", min_value=0.0, step=0.01,
                              key=f"p_mix_{i}", label_visibility="collapsed")
    with cols[6]:
        il = st.number_input("Ηλ/γος", min_value=0.0, step=0.01,
                             key=f"p_il_{i}", label_visibility="collapsed")
    parts.append({'name':name,'type':ptype,'price':price,'fanop':fanop,
                  'vafeas':vafeas,'mixanikos':mix,'ilgos':il})

    # Hint τιμής ανά τύπο αν υπάρχει ιστορικό
    if name and ptype and marka and montelo:
        stats_all = get_antallaktiko_stats_all_types(marka, montelo, name)
        if stats_all:
            hints = []
            for t, s in stats_all.items():
                marker = " ◀" if t == ptype else ""
                hints.append(f"**{t}**: {s['mean']:.0f}€ ({s['count']} φορές){marker}")
            st.caption("💡 Ιστορικό τιμών: " + " | ".join(hints))
        elif name and len(name) > 3 and marka and montelo:
            pass  # Δεν εμφανίζουμε τίποτα αν δεν υπάρχουν δεδομένα

pb1, pb2, _ = st.columns([1,1,4])
with pb1:
    if st.button("➕ Προσθήκη γραμμής", key="add_part", use_container_width=True):
        st.session_state.num_parts += 1; st.rerun()
with pb2:
    if st.button("➖ Αφαίρεση γραμμής", key="remove_part", use_container_width=True,
                 disabled=st.session_state.num_parts <= 1):
        st.session_state.num_parts -= 1; st.rerun()
st.markdown("---")

# ============================================================
# ΕΡΓΑΣΙΕΣ
# ============================================================
st.subheader("⚙️ Εργασίες")

works = []
hc = st.columns([2,3,1,1,1,1])
for col, lbl in zip(hc, ["**Τύπος**","**Περιγραφή**","**Φανοπ.**","**Βαφέας**","**Μηχ/κός**","**Ηλ/γος**"]):
    col.markdown(lbl)

for i in range(st.session_state.num_works):
    cols = st.columns([2,3,1,1,1,1])
    with cols[0]:
        wtype = st.text_input("Τύπος εργασίας",
                             key=f"w_type_{i}", label_visibility="collapsed",
                             placeholder="ΕΠΙΣΚΕΥΗ")
    with cols[1]:
        desc = st.text_input("Περιγραφή", key=f"w_desc_{i}",
                             label_visibility="collapsed", placeholder="π.χ. ΠΡΟΦΥΛΑΚΤΗΡΑΣ ΠΙΣΩ")
    with cols[2]:
        fanop = st.number_input("Φανοπ.", min_value=0.0, step=0.01,
                                key=f"w_fanop_{i}", label_visibility="collapsed")
    with cols[3]:
        vafeas = st.number_input("Βαφέας", min_value=0.0, step=0.01,
                                 key=f"w_vafeas_{i}", label_visibility="collapsed")
    with cols[4]:
        mix = st.number_input("Μηχ/κός", min_value=0.0, step=0.01,
                              key=f"w_mix_{i}", label_visibility="collapsed")
    with cols[5]:
        il = st.number_input("Ηλ/γος", min_value=0.0, step=0.01,
                             key=f"w_il_{i}", label_visibility="collapsed")
    works.append({'type':wtype,'desc':desc,'price':0,'fanop':fanop,
                  'vafeas':vafeas,'mixanikos':mix,'ilgos':il})

wb1, wb2, _ = st.columns([1,1,4])
with wb1:
    if st.button("➕ Προσθήκη γραμμής", key="add_work", use_container_width=True):
        st.session_state.num_works += 1; st.rerun()
with wb2:
    if st.button("➖ Αφαίρεση γραμμής", key="remove_work", use_container_width=True,
                 disabled=st.session_state.num_works <= 1):
        st.session_state.num_works -= 1; st.rerun()

# ============================================================
# ΣΥΝΟΛΑ
# ============================================================
st.markdown("---")
st.subheader("💰 Σύνολα (live preview)")
subtotal = sum(r['price']+r['fanop']+r['vafeas']+r['mixanikos']+r['ilgos'] for r in parts+works)
vat   = subtotal * 0.24
total = subtotal + vat
c1,c2,c3 = st.columns(3)
c1.metric("Σύνολο χωρίς ΦΠΑ", f"{subtotal:.2f} €")
c2.metric("ΦΠΑ 24%", f"{vat:.2f} €")
c3.metric("Γενικό Σύνολο", f"{total:.2f} €")
st.markdown("---")

# ============================================================
# ΠΑΡΑΤΗΡΗΣΕΙΣ
# ============================================================
st.subheader("📝 Παρατηρήσεις")
template_checks = {k: st.checkbox(k, key=f"tmpl_{k}") for k in PARATIRISI_TEMPLATES}
default_text = "\n\n".join(PARATIRISI_TEMPLATES[k] for k,v in template_checks.items() if v)
paratiriseis = st.text_area("Κείμενο παρατηρήσεων (επεξεργάσιμο)",
                             value=default_text, height=200,
                             key=f"par_{hash(default_text)}")
st.markdown("---")

# ============================================================
# ΦΩΤΟΓΡΑΦΙΕΣ (2 ανά σελίδα)
# ============================================================
st.subheader("📸 Φωτογραφικό Υλικό")
st.caption("2 φωτογραφίες ανά σελίδα — μεγάλες.")

# Screenshot car.gr
screenshot_file = st.file_uploader(
    "📱 Screenshot αγγελιών car.gr (προαιρετικό)",
    type=['jpg','jpeg','png'],
    key="screenshot_cargr",
    help="Κάνε screenshot από το car.gr και επισύναψέ το εδώ"
)
if screenshot_file:
    st.image(screenshot_file, caption="Screenshot car.gr", use_container_width=True)

st.markdown("**Φωτογραφίες ζημιάς:**")
uploaded_files = st.file_uploader("Επιλέξτε φωτογραφίες (JPG/PNG)",
                                   type=['jpg','jpeg','png'], accept_multiple_files=True)
photo_captions = []
if uploaded_files:
    st.write(f"**{len(uploaded_files)} φωτογραφίες:**")
    for i, file in enumerate(uploaded_files):
        c1, c2 = st.columns([2,1])
        with c1:
            st.image(file, use_container_width=True)
        with c2:
            cap = st.text_input(f"Περιγραφή #{i+1}", key=f"caption_{i}",
                                placeholder="π.χ. Πίσω προφυλακτήρας")
            photo_captions.append(cap)
        if i < len(uploaded_files)-1:
            st.markdown("---")
st.markdown("---")

# ============================================================
# ΟΝΟΜΑΤΕΠΩΝΥΜΟ
# ============================================================
onomateponymo = st.text_input("Ονοματεπώνυμο Πραγματογνώμονα")
st.markdown("---")


# ============================================================
# ΒΟΗΘΗΤΙΚΕΣ ΣΥΝΑΡΤΗΣΕΙΣ
# ============================================================
def fmt_date(d):
    return d.strftime("%d/%m/%Y") if d else ""

def fmt_part(p):
    if not p['name']: return ""
    r = f"ΑΝΤΙΚΑΤΑΣΤΑΣΗ {p['name']}"
    if p['type']: r += f" ({p['type']})"
    return r

def fmt_work(w):
    if not w['type'] and not w['desc']: return ""
    return (w['type'] + " " + w['desc']).strip()

def compress_image(image_bytes, max_width=1600, quality=82):
    img = PILImage.open(io.BytesIO(image_bytes))
    if img.mode in ('RGBA','P'): img = img.convert('RGB')
    if img.width > max_width:
        img = img.resize((max_width, int(img.height*max_width/img.width)),
                         PILImage.Resampling.LANCZOS)
    out = io.BytesIO()
    img.save(out, format='JPEG', quality=quality, optimize=True)
    out.seek(0)
    return out

def fill_excel(template_path, _uploaded_files=None, _photo_captions=None):
    wb = load_workbook(template_path)
    ws = wb.active

    # Λογότυπα ανάλογα με ασφαλιστική
    asfalistiki_sel = st.session_state.get('asfalistiki', 'INTERLIFE')
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    if asfalistiki_sel == 'INTERLIFE':
        # INTERLIFE: κρατάμε τα αρχικά λογότυπα GNOMON + TUV, μεγαλώνουμε
        for img in ws._images:
            try:
                a = img.anchor
                col = a._from.col
                if col == 1:
                    img.width  = 1100
                    img.height = 400
                elif col == 6:
                    img.width  = 420
                    img.height = 420
            except Exception:
                pass
    else:
        # ΕΘΝΙΚΗ / APEIRON: αφαιρούμε τα GNOMON/TUV, βάζουμε experts360
        LOGO_PATH = "logo_experts360.png"
        ws._images.clear()
        if os.path.exists(LOGO_PATH):
            logo = XLImage(LOGO_PATH)
            logo.width  = 220
            logo.height = 75
            marker = AnchorMarker(col=1, colOff=pixels_to_EMU(5),
                                  row=0, rowOff=pixels_to_EMU(5))
            size   = XDRPositiveSize2D(pixels_to_EMU(220), pixels_to_EMU(75))
            logo.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(logo)

    ws['C5'] = ar_zimias
    ws['E5'] = hm_entolhs or ''
    ws['C6'] = hm_atyxhmatos or ''
    ws['C7'] = topos_atyxhmatos
    ws['C10'] = idioktitis
    ws['C11'] = ar_kykloforias
    ws['E11'] = xrisi
    ws['C12'] = marka
    ws['E12'] = kyvika
    ws['G12'] = montelo
    ws['C13'] = proti_adeia
    ws['E13'] = fmt_date(hm_kteo)
    ws['C14'] = xiliometrites
    ws['E14'] = f"{axia} €" if axia > 0 else ""
    # Αρ. Πλαισίου VIN - γραμμή 15 (παρατηρήσεις κελί)
    if ar_plaisiou:
        from openpyxl.styles import Font as XLFont, Alignment as XLAlign
        ws['B15'] = 'Αρ. Πλαισίου (VIN):'
        ws['B15'].font = XLFont(bold=True, size=10)
        ws['C15'] = ar_plaisiou
        ws['C15'].font = XLFont(size=10)
        ws['C15'].alignment = XLAlign(horizontal='left')
        ws.row_dimensions[15].height = 18

    # Επιπλέον πεδία για ΕΘΝΙΚΗ/APEIRON - μόνο αν συμπληρωμένα
    asfalistiki_xl = st.session_state.get('asfalistiki','INTERLIFE')
    if asfalistiki_xl in ('ΕΘΝΙΚΗ ΑΣΦΑΛΙΣΤΙΚΗ','APEIRON ΑΣΦΑΛΙΣΤΙΚΗ'):
        from openpyxl.styles import Font as XLFont2, Alignment as XLAlign2, PatternFill
        extra_row = 16
        extra_fields_xl = [
            ('xroma', 'Χρώμα:', 'kaysimo', 'Καύσιμο:'),
            ('katast_oxima', 'Κατάσταση Οχήματος:', 'kindynos', 'Κίνδυνος:'),
            ('ixni_xromatos', 'Ίχνη Χρωμάτων:', 'fora_atyxima', 'Φορά Ατυχήματος:'),
            ('elastikon_simeio', 'Σημείο Ελαστικών:', 'tilefono', 'Τηλέφωνο Ιδιοκτήτη:'),
        ]
        # Αποθηκεύω τιμές στο session state
        _extra_vals = {
            'xroma': st.session_state.get('xroma',''),
            'kaysimo': st.session_state.get('kaysimo',''),
            'katast_oxima': st.session_state.get('katast_oxima',''),
            'kindynos': st.session_state.get('kindynos',''),
            'ixni_xromatos': st.session_state.get('ixni_xromatos',''),
            'fora_atyxima': st.session_state.get('fora_atyxima',''),
            'elastikon_simeio': st.session_state.get('elastikon_simeio',''),
            'tilefono': st.session_state.get('tilefono',''),
        }
        for k1, l1, k2, l2 in extra_fields_xl:
            v1 = _extra_vals.get(k1,'')
            v2 = _extra_vals.get(k2,'')
            if v1 or v2:
                ws.insert_rows(extra_row)
                ws[f'B{extra_row}'] = l1
                ws[f'B{extra_row}'].font = XLFont2(bold=True, size=9)
                ws[f'C{extra_row}'] = v1
                ws[f'D{extra_row}'] = l2
                ws[f'D{extra_row}'].font = XLFont2(bold=True, size=9)
                ws[f'E{extra_row}'] = v2
                ws.row_dimensions[extra_row].height = 16
                extra_row += 1

        # Συνεργείο
        syn_vals = {
            'eponimia': st.session_state.get('synergeio_eponimia',''),
            'dieuthinsi': st.session_state.get('synergeio_dieuthinsi',''),
            'tilefono': st.session_state.get('synergeio_tilefono',''),
            'kinito': st.session_state.get('synergeio_kinito',''),
            'fax': st.session_state.get('synergeio_fax',''),
            'mail': st.session_state.get('synergeio_mail',''),
        }
        if any(syn_vals.values()):
            ws.insert_rows(extra_row)
            ws[f'B{extra_row}'] = 'ΣΤΟΙΧΕΙΑ ΣΥΝΕΡΓΕΙΟΥ'
            ws[f'B{extra_row}'].font = XLFont2(bold=True, size=10)
            extra_row += 1
            for lbl, val in [
                ('Επωνυμία:', syn_vals['eponimia']),
                ('Διεύθυνση:', syn_vals['dieuthinsi']),
                ('Τηλέφωνο:', syn_vals['tilefono']),
                ('Κινητό:', syn_vals['kinito']),
                ('Fax:', syn_vals['fax']),
                ('Email:', syn_vals['mail']),
            ]:
                if val:
                    ws.insert_rows(extra_row)
                    ws[f'B{extra_row}'] = lbl
                    ws[f'B{extra_row}'].font = XLFont2(bold=True, size=9)
                    ws[f'C{extra_row}'] = val
                    ws.row_dimensions[extra_row].height = 16
                    extra_row += 1

    for i, v in enumerate(visits[:2]):
        row = 17 + i
        ds = v.get('date_text','')
        ps = v.get('place','')
        if i == 0:
            ws[f'C{row}'] = ds
            ws[f'E{row}'] = ps
        else:
            ws[f'C{row}'] = f"{ds} - {ps}" if ps else ds

    extra = ""
    for i, v in enumerate(visits[2:], start=3):
        ds = v.get('date_text','')
        if ds or v['place']:
            extra += f"\nΗμ/νία {i}ης Επίσκεψης: {ds}"
            if v['place']: extra += f" - {v['place']}"

    for i, p in enumerate(parts[:10]):
        row = 21+i
        d = fmt_part(p)
        if d:
            ws[f'B{row}'] = d
            if p['price']>0:    ws[f'C{row}'] = p['price']
            if p['fanop']>0:    ws[f'D{row}'] = p['fanop']
            if p['vafeas']>0:   ws[f'E{row}'] = p['vafeas']
            if p['mixanikos']>0:ws[f'F{row}'] = p['mixanikos']
            if p['ilgos']>0:    ws[f'G{row}'] = p['ilgos']

    for i, w in enumerate(works[:10]):
        row = 33+i
        d = fmt_work(w)
        if d:
            ws[f'B{row}'] = d
            if w['fanop']>0:    ws[f'D{row}'] = w['fanop']
            if w['vafeas']>0:   ws[f'E{row}'] = w['vafeas']
            if w['mixanikos']>0:ws[f'F{row}'] = w['mixanikos']
            if w['ilgos']>0:    ws[f'G{row}'] = w['ilgos']

    fp = paratiriseis
    if extra: fp = "Επιπλέον επισκέψεις:" + extra + "\n\n" + fp
    ws['B51'] = fp
    # Επιπλέον πεδία για ΕΘΝΙΚΗ/APEIRON (μόνο αν έχουν τιμή)
    if asfalistiki_sel in ('ΕΘΝΙΚΗ ΑΣΦΑΛΙΣΤΙΚΗ', 'APEIRON ΑΣΦΑΛΙΣΤΙΚΗ'):
        from openpyxl.styles import Font as XLFont2
        extra_row = 16
        extra_fields = [
            ('Τηλέφωνο:', tilefono),
            ('Χρώμα:', xroma),
            ('Καύσιμο:', kaysimo),
            ('Κατάσταση Οχήματος:', katast_oxima),
            ('Ίχνη Χρωμάτων:', ixni_xromatos),
            ('Φορά Ατυχήματος:', fora_atyxima),
            ('Σημείο Ελαστικών:', elastikon_simeio),
            ('Κίνδυνος:', kindynos),
            ('Συνεργείο:', synergeio_eponimia),
            ('Διεύθυνση:', synergeio_dieuthinsi),
            ('Τηλ. Συνεργείου:', synergeio_tilefono),
            ('Κινητό:', synergeio_kinito),
            ('Fax:', synergeio_fax),
            ('Email:', synergeio_mail),
        ]
        for lbl, val in extra_fields:
            if val and str(val).strip():
                ws[f'B{extra_row}'] = lbl
                ws[f'B{extra_row}'].font = XLFont2(bold=True, size=9)
                ws[f'C{extra_row}'] = str(val)
                ws[f'C{extra_row}'].font = XLFont2(size=9)
                ws.row_dimensions[extra_row].height = 16
                extra_row += 1

    asfalistiki_sign = st.session_state.get('asfalistiki', 'INTERLIFE')
    if asfalistiki_sign == 'INTERLIFE':
        ws['B61'] = 'ΓΙΑ ΤΗΝ GNOMON EXPERTS Α.Ε.'
    else:
        ws['B61'] = 'ΓΙΑ ΤΗΝ EXPERTS360'
    ws['B62'] = onomateponymo

    # Φωτογραφίες 2 ανά σελίδα
    screenshot = st.session_state.get('screenshot_cargr')
    all_files = []
    all_captions = list(_photo_captions or [])
    if screenshot:
        all_files.insert(0, screenshot)
        all_captions.insert(0, "Αγγελίες car.gr - Εμπορική αξία")
    if _uploaded_files:
        all_files.extend(_uploaded_files)

    if all_files:
        from openpyxl.worksheet.pagebreak import Break
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

        start_row = 65
        ws.row_dimensions[start_row].height = 24
        ws[f'B{start_row}'] = 'ΦΩΤΟΓΡΑΦΙΚΟ ΥΛΙΚΟ'
        ws[f'B{start_row}'].font = Font(name='Calibri', size=14, bold=True)
        ws[f'B{start_row}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'B{start_row}:H{start_row}')
        ws.row_breaks.append(Break(id=63))

        # Page setup για σωστή εκτύπωση φωτογραφιών
        ws.page_setup.paperSize   = 9   # A4
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.scale       = 100
        # Print area θα οριστεί στο τέλος

        # Διαστάσεις φωτό σε pixels (A4 ~700px πλάτος χρήσιμο)
        IMG_W = 480  # pixels
        IMG_H = 340  # pixels

        # Κάθε block: 20 γραμμές φωτό + 2 λεζάντα + 1 κενή = 23
        photo_h   = 20
        cap_rows  = 2
        gap       = 1
        block     = photo_h + cap_rows + gap   # 23 γραμμές
        page_rows = block * 2 + 2              # 48 γραμμές ανά σελίδα

        page_start = start_row + 2

        # Ύψος γραμμής ~15pt -> ~20px. Για κεντράρισμα:
        # Το template έχει στήλες B-H. Συνολικό πλάτος ~700px
        # Κεντράρουμε βάζοντας offset ~110px από αριστερά (col B)
        LEFT_OFFSET_PX = 105  # pixels από αρχή col B για κεντράρισμα

        for idx, file in enumerate(all_files):
            pos = idx % 2
            photo_row = page_start if pos == 0 else page_start + block

            file.seek(0)
            comp = compress_image(file.read(), max_width=1400, quality=84)
            tmp = f'/tmp/photo_{idx}.jpg'
            with open(tmp, 'wb') as f:
                f.write(comp.getvalue())

            img = XLImage(tmp)
            img.width  = IMG_W
            img.height = IMG_H

            # Anchor: col B (index 1), row=photo_row-1 (0-based)
            # colOff σε EMU για οριζόντιο κεντράρισμα
            col_off_emu = pixels_to_EMU(LEFT_OFFSET_PX)
            row_off_emu = pixels_to_EMU(2)  # μικρό padding πάνω

            marker = AnchorMarker(col=1, colOff=col_off_emu,
                                  row=photo_row-1, rowOff=row_off_emu)
            size   = XDRPositiveSize2D(pixels_to_EMU(IMG_W), pixels_to_EMU(IMG_H))
            anchor = OneCellAnchor(_from=marker, ext=size)
            img.anchor = anchor
            ws.add_image(img)

            # Λεζάντα: merged κελί B:H, 2 γραμμές κάτω από τη φωτό
            cap_row = photo_row + photo_h
            cap_txt = photo_captions[idx] if idx < len(photo_captions) else ''
            ws[f'B{cap_row}'] = cap_txt
            ws[f'B{cap_row}'].font = Font(name='Calibri', size=11, italic=True)
            ws[f'B{cap_row}'].alignment = Alignment(horizontal='center',
                                                     vertical='center',
                                                     wrap_text=True)
            try:
                ws.merge_cells(f'B{cap_row}:H{cap_row}')
            except Exception:
                pass
            ws.row_dimensions[cap_row].height = 30

            # Page break μετά 2η φωτό
            if pos == 1 and idx < len(all_files) - 1:
                ws.row_breaks.append(Break(id=page_start + page_rows - 1))
                page_start += page_rows

    # Ορίζουμε print area που καλύπτει ΟΛΟ το περιεχόμενο (έκθεση + φωτό)
    last_row = ws.max_row + 10
    ws.print_area = f'$A$1:$H${last_row}'

    # Περιθώρια και page setup
    ws.page_setup.paperSize   = 9        # A4
    ws.page_setup.orientation = 'portrait'
    ws.page_margins.left      = 0.4
    ws.page_margins.right     = 0.4
    ws.page_margins.top       = 0.4
    ws.page_margins.bottom    = 0.4

    return wb

def excel_to_pdf(xlsx_bytes):
    with tempfile.TemporaryDirectory() as d:
        xp = os.path.join(d,"ekthesi.xlsx")
        with open(xp,'wb') as f: f.write(xlsx_bytes)
        subprocess.run(['libreoffice','--headless','--convert-to','pdf','--outdir',d,xp],
                       capture_output=True, timeout=60)
        pp = os.path.join(d,"ekthesi.pdf")
        if os.path.exists(pp):
            with open(pp,'rb') as f: return f.read()
        return None

# ============================================================
# ΕΞΑΓΩΓΗ
# ============================================================
# ============================================================
# ΑΠΟΘΗΚΕΥΣΗ ΣΤΗ ΒΑΣΗ
# ============================================================
st.subheader("💾 Αποθήκευση")

sa_col1, sa_col2, sa_col3 = st.columns([2,2,2])
with sa_col1:
    save_user = st.text_input("Όνομα χρήστη", value=st.session_state.get('save_user',''),
                               placeholder="π.χ. Μαρινάκης", key="save_user")
with sa_col2:
    save_status = st.selectbox("Status",["draft","final","archived"],
                               format_func=lambda x:{"draft":"Προσχέδιο","final":"Τελική","archived":"Αρχείο"}[x],
                               key="save_status")
with sa_col3:
    st.write("")
    st.write("")
    eid = st.session_state.get('current_ekthesi_id')
    btn_label = f"💾 Ενημέρωση #{eid}" if eid else "💾 Αποθήκευση νέας"
    if st.button(btn_label, type="primary", use_container_width=True):
        save_data = {
            "ar_zimias":       ar_zimias,
            "hm_entolhs":      str(hm_entolhs) if hm_entolhs else "",
            "hm_atyxhmatos":   str(hm_atyxhmatos) if hm_atyxhmatos else "",
            "topos_atyxhmatos":topos_atyxhmatos,
            "idioktitis":      idioktitis,
            "ar_kykloforias":  ar_kykloforias,
            "ar_plaisiou":     ar_plaisiou,
            "marka":           marka,
            "montelo":         montelo,
            "kyvika":          kyvika,
            "xrisi":           xrisi,
            "proti_adeia":     proti_adeia,
            "xiliometrites":   xiliometrites,
            "axia":            axia,
            "hm_kteo":         str(hm_kteo) if hm_kteo else "",
            "visit_date":      visits[0]['date_text'] if visits else "",
            "visit_place":     visits[0]['place'] if visits else "",
            "paratiriseis":    paratiriseis,
            "onomateponymo":   onomateponymo,
            "status":          save_status,
        }
        new_id, err = save_ekthesi(
            data=save_data, parts=parts, works=works,
            user_name=save_user, ekthesi_id=eid
        )
        if err:
            st.error(f"❌ Σφάλμα αποθήκευσης: {err}")
        else:
            st.session_state['current_ekthesi_id'] = new_id
            st.success(f"✅ Αποθηκεύτηκε με ID #{new_id}")

st.markdown("---")
st.subheader("📤 Εξαγωγή Έκθεσης")

if not os.path.exists(TEMPLATE_FILE):
    st.error(f"⚠️ Δεν βρέθηκε το αρχείο: {TEMPLATE_FILE}")
else:
    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("📊 Δημιουργία Excel", type="primary", use_container_width=True):
            try:
                _ufiles = uploaded_files if 'uploaded_files' in dir() and uploaded_files else []
                _pcaps  = photo_captions if 'photo_captions' in dir() and photo_captions else []
                wb = fill_excel(TEMPLATE_FILE, _uploaded_files=_ufiles, _photo_captions=_pcaps)
                out = io.BytesIO(); wb.save(out); out.seek(0)
                fn = f"ekthesi_{ar_zimias.replace('/','_') if ar_zimias else 'nea'}.xlsx"
                st.download_button("⬇️ Κατέβασμα Excel", data=out, file_name=fn,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
            except Exception as e:
                st.error(f"Σφάλμα Excel: {e}"); st.exception(e)
    with c2:
        if st.button("📄 Δημιουργία PDF", type="primary", use_container_width=True):
            try:
                with st.spinner("Δημιουργία PDF..."):
                    _ar_plaisiou = st.session_state.get('ar_plaisiou','')
                    pdf_data = {
                        'ar_zimias':        ar_zimias,
                        'hm_entolhs':       str(hm_entolhs) if hm_entolhs else '',
                        'hm_atyxhmatos':    str(hm_atyxhmatos) if hm_atyxhmatos else '',
                        'topos_atyxhmatos': topos_atyxhmatos,
                        'idioktitis':       idioktitis,
                        'ar_kykloforias':   ar_kykloforias,
                        'ar_plaisiou':      _ar_plaisiou,
                        'marka':            marka,
                        'montelo':          montelo,
                        'kyvika':           kyvika,
                        'xrisi':            xrisi,
                        'proti_adeia':      proti_adeia,
                        'xiliometrites':    xiliometrites,
                        'axia':             axia,
                        'hm_kteo':          str(hm_kteo) if hm_kteo else '',
                        'tilefono':         tilefono,
                        'xroma':            xroma,
                        'kaysimo':          kaysimo,
                        'katast_oxima':     katast_oxima,
                        'ixni_xromatos':    ixni_xromatos,
                        'fora_atyxima':     fora_atyxima,
                        'elastikon_simeio': elastikon_simeio,
                        'kindynos':         kindynos,
                        'synergeio_eponimia':   synergeio_eponimia,
                        'synergeio_dieuthinsi': synergeio_dieuthinsi,
                        'synergeio_tilefono':   synergeio_tilefono,
                        'synergeio_kinito':     synergeio_kinito,
                        'synergeio_fax':        synergeio_fax,
                        'synergeio_mail':       synergeio_mail,
                    }
                    asfalistiki_sel = st.session_state.get('asfalistiki','INTERLIFE')
                    pdf = generate_pdf(
                        data=pdf_data, parts=parts, works=works,
                        visits=visits,
                        photo_files=uploaded_files if uploaded_files else [],
                        photo_captions=photo_captions,
                        paratiriseis=paratiriseis,
                        onomateponymo=onomateponymo,
                        asfalistiki=asfalistiki_sel,
                    )
                fn = f"ekthesi_{ar_zimias.replace('/','_') if ar_zimias else 'nea'}.pdf"
                st.success(f"✓ PDF έτοιμο ({len(pdf)/1024:.0f} KB)")
                st.download_button("⬇️ Κατέβασμα PDF", data=pdf, file_name=fn,
                                   mime="application/pdf", use_container_width=True)
            except Exception as e:
                st.error(f"Σφάλμα PDF: {e}"); st.exception(e)
    with c3:
        if st.button("☁️ Αποθήκευση στο Drive", use_container_width=True):
            try:
                import sys, os
                sys.path.insert(0, os.path.dirname(__file__))
                from gdrive import upload_ekthesi_files, get_ekthesi_folder_link
                with st.spinner("Ανέβασμα στο Google Drive..."):
                    drive_files = {}
                    # PDF
                    _pdf_data2 = {
                        'ar_zimias': ar_zimias, 'hm_entolhs': str(hm_entolhs) if hm_entolhs else '',
                        'hm_atyxhmatos': str(hm_atyxhmatos) if hm_atyxhmatos else '',
                        'topos_atyxhmatos': topos_atyxhmatos, 'idioktitis': idioktitis,
                        'ar_kykloforias': ar_kykloforias, 'ar_plaisiou': st.session_state.get('ar_plaisiou',''),
                        'marka': marka, 'montelo': montelo, 'kyvika': kyvika, 'xrisi': xrisi,
                        'proti_adeia': proti_adeia, 'xiliometrites': xiliometrites,
                        'axia': axia, 'hm_kteo': str(hm_kteo) if hm_kteo else '',
                    }
                    _pdf_bytes = generate_pdf(
                        data=_pdf_data2, parts=parts, works=works, visits=visits,
                        photo_files=[], photo_captions=[], paratiriseis=paratiriseis,
                        onomateponymo=onomateponymo, asfalistiki=st.session_state.get('asfalistiki','INTERLIFE')
                    )
                    drive_files['pdf'] = _pdf_bytes
                    # Excel
                    _ufiles2 = uploaded_files if 'uploaded_files' in dir() and uploaded_files else []
                    _pcaps2  = photo_captions if 'photo_captions' in dir() and photo_captions else []
                    _wb = fill_excel(TEMPLATE_FILE, _uploaded_files=_ufiles2, _photo_captions=_pcaps2)
                    _xls_buf = io.BytesIO(); _wb.save(_xls_buf); _xls_buf.seek(0)
                    drive_files['excel'] = _xls_buf.read()
                    # Φωτογραφίες
                    if _ufiles2:
                        drive_files['photos'] = []
                        for f2 in _ufiles2:
                            f2.seek(0)
                            drive_files['photos'].append((f2.read(), f2.name))
                    links = upload_ekthesi_files(ar_zimias, drive_files)
                if links:
                    st.success("✅ Αποθηκεύτηκε στο Drive!")
                    folder_link = get_ekthesi_folder_link(ar_zimias)
                    if folder_link:
                        st.markdown(f"[📁 Άνοιγμα φακέλου]({folder_link})")
                else:
                    st.error("❌ Αποτυχία ανεβάσματος")
            except Exception as e:
                st.error(f"Drive error: {e}")
